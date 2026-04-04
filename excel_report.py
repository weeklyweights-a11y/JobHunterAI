"""Excel report generation (Phase 5) — styled .xlsx with apply hyperlinks."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from output_util import OUTPUT_DIR

logger = logging.getLogger(__name__)

_FONT_NAME = "Calibri"
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_ALT_ROW_FILLS = (
    PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
)
_MAX_COL_WIDTH = 60


def _source_label(key: str) -> str:
    m = {
        "linkedin": "LinkedIn",
        "indeed": "Indeed",
        "ats": "ATS",
        "greenhouse": "Greenhouse",
        "yc": "YC",
        "career_page": "Career Pages",
    }
    return m.get(key, key or "Unknown")


def _apply_type_label(t: str) -> str:
    t = (t or "").strip().lower()
    if t == "easy_apply":
        return "Easy Apply"
    if t == "external":
        return "External"
    return t or ""


def _header_font() -> Font:
    return Font(name=_FONT_NAME, size=11, bold=True, color="000000")


def _body_font() -> Font:
    return Font(name=_FONT_NAME, size=11, color="000000")


def _title_font() -> Font:
    return Font(name=_FONT_NAME, size=14, bold=True, color="000000")


def _prepare_rows(
    jobs: list[dict[str, Any]],
) -> list[tuple[str, ...]]:
    """Title, Company, Location, Seniority, Freshness, Posted Time, Applicant Count, Apply Type, Source, Job ID, Apply Link, Found At. (No job description — LLM-only.)"""
    rows: list[tuple[str, ...]] = []
    for j in jobs:
        src = _source_label(str(j.get("source") or ""))
        found = str(j.get("found_at") or "")
        rows.append(
            (
                str(j.get("title") or ""),
                str(j.get("company") or ""),
                str(j.get("location") or ""),
                str(j.get("seniority") or ""),
                str(j.get("freshness") or ""),
                str(j.get("posted_time") or ""),
                str(j.get("applicant_count") or ""),
                _apply_type_label(str(j.get("apply_type") or "")),
                src,
                str(j.get("job_id") or ""),
                str(j.get("url") or ""),
                found,
            )
        )
    return rows


def _autosize_columns(ws: Any, headers: list[str], data_rows: list[tuple]) -> None:
    widths: list[int] = [len(h) for h in headers]
    for row in data_rows:
        for i, cell in enumerate(row):
            if i < len(widths):
                widths[i] = max(widths[i], len(cell))
    for col_idx, w in enumerate(widths, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(_MAX_COL_WIDTH, w + 2)


def write_jobs_xlsx(
    jobs: list[dict[str, Any]],
    *,
    output_dir: Path | None = None,
) -> Path:
    """Write Phase 5 styled workbook; return path. Uses local time for filename."""
    out = output_dir or OUTPUT_DIR
    out.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    path = out / f"jobs_{stamp}.xlsx"
    date_title = datetime.now().strftime("%Y-%m-%d")

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [
        "Job Title",
        "Company",
        "Location",
        "Seniority",
        "Freshness",
        "Posted Time",
        "Applicant Count",
        "Apply Type",
        "Source",
        "Job ID",
        "Apply Link",
        "Found At",
    ]
    ncols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c1 = ws.cell(row=1, column=1)
    c1.value = f"JobHunter AI — Jobs Found on {date_title}"
    c1.font = _title_font()
    c1.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=1).value = None

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = _header_font()
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_rows = _prepare_rows(jobs)
    apply_col = 11
    for i, row_tuple in enumerate(data_rows):
        r = 4 + i
        fill = _ALT_ROW_FILLS[i % 2]
        for c_idx, val in enumerate(row_tuple, start=1):
            if c_idx == apply_col:
                cell = ws.cell(row=r, column=c_idx)
                url = val
                if url:
                    cell.hyperlink = url
                    cell.value = "Apply →"
                    cell.font = Font(
                        name=_FONT_NAME,
                        size=11,
                        color="0563C1",
                        underline="single",
                    )
                else:
                    cell.value = ""
                    cell.font = _body_font()
            else:
                cell = ws.cell(row=r, column=c_idx, value=val)
                cell.font = _body_font()
                # Force text so Excel does not parse ISO-8601 as broken date/time (e.g. ":51-04:00").
                if c_idx in (6, 12):
                    cell.number_format = "@"
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary_row = 4 + len(data_rows)
    ws.cell(row=summary_row, column=1, value=f"Total: {len(jobs)} new jobs")
    ws.cell(row=summary_row, column=1).font = Font(
        name=_FONT_NAME, size=11, bold=True, color="000000"
    )

    _autosize_columns(ws, headers, data_rows)
    ws.freeze_panes = "A4"

    try:
        wb.save(path)
    except OSError:
        logger.exception("Failed to save Excel to %s", path)
        raise
    logger.info("Wrote Excel report: %s", path)
    return path
